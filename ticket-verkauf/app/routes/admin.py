import io
from datetime import timezone
from functools import wraps

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import (
    Blueprint,
    Response,
    abort,
    redirect,
    render_template,
    request,
    session,
    url_for,
)

from ..db import (
    get_all_registrations,
    get_settings,
    get_ticket_counts,
    set_setting,
    verify_admin,
)

bp = Blueprint("admin", __name__)

EDITABLE_SETTINGS = [
    "ticket_price",
    "total_tickets",
    "paypal_name",
    "paypal_email",
    "bank_owner",
    "bank_iban",
    "bank_bic",
    "bank_reference",
]


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin.login"))
        return f(*args, **kwargs)
    return decorated


@bp.route("/login", methods=["GET", "POST"])
def login():
    if session.get("admin_logged_in"):
        return redirect(url_for("admin.dashboard"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if verify_admin(username, password):
            session.permanent = False
            session["admin_logged_in"] = True
            session["admin_username"] = username
            return redirect(url_for("admin.dashboard"))
        error = "Ungültige Anmeldedaten."

    return render_template("admin/login.html", error=error)


@bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("admin.login"))


@bp.route("/")
@admin_required
def dashboard():
    settings = get_settings()
    registrations = get_all_registrations()
    total, sold, remaining = get_ticket_counts()
    return render_template(
        "admin/dashboard.html",
        settings=settings,
        registrations=registrations,
        total=total,
        sold=sold,
        remaining=remaining,
    )


@bp.route("/toggle-registration", methods=["POST"])
@admin_required
def toggle_registration():
    current = get_settings().get("registration_open", "false")
    set_setting("registration_open", "false" if current.lower() == "true" else "true")
    return redirect(url_for("admin.dashboard"))


@bp.route("/toggle-fcfs", methods=["POST"])
@admin_required
def toggle_fcfs():
    current = get_settings().get("fcfs_block_enabled", "false")
    set_setting("fcfs_block_enabled", "false" if current.lower() == "true" else "true")
    return redirect(url_for("admin.dashboard"))


@bp.route("/toggle-payment/<int:reg_id>", methods=["POST"])
@admin_required
def toggle_payment(reg_id: int):
    from ..db import get_conn
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE registrations SET payment_confirmed = NOT payment_confirmed WHERE id = %s",
                (reg_id,),
            )
    return redirect(url_for("admin.dashboard"))


@bp.route("/delete-registration/<int:reg_id>", methods=["POST"])
@admin_required
def delete_registration(reg_id: int):
    from ..db import get_conn
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM registrations WHERE id = %s", (reg_id,))
    return redirect(url_for("admin.dashboard"))


@bp.route("/delete-all-registrations", methods=["POST"])
@admin_required
def delete_all_registrations():
    from ..db import get_conn
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM registrations")
    return redirect(url_for("admin.dashboard"))


@bp.route("/update-settings", methods=["POST"])
@admin_required
def update_settings():
    for key in EDITABLE_SETTINGS:
        value = request.form.get(key, "").strip()
        if value != "":
            set_setting(key, value)
    return redirect(url_for("admin.dashboard"))


@bp.route("/export")
@admin_required
def export():
    registrations = get_all_registrations()
    settings = get_settings()
    price = int(settings.get("ticket_price", 60))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrierungen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3a5c")
    center = Alignment(horizontal="center")

    headers = [
        "Reg-Nr.",
        "Zeitstempel (UTC)",
        "Hauptperson",
        "E-Mail",
        "Wunschtisch",
        "Tickets gesamt",
        "Betrag (€)",
        "Zahlung bestätigt",
        "Person",
        "Über 18",
        "Essen",
        "Typ",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = 2
    for reg in registrations:
        ts = reg["created_at"]
        if hasattr(ts, "astimezone"):
            ts = ts.astimezone(timezone.utc).strftime("%d.%m.%Y %H:%M:%S")

        for person in reg["persons"]:
            ws.cell(row=row, column=1, value=reg["id"])
            ws.cell(row=row, column=2, value=str(ts))
            ws.cell(row=row, column=3, value=reg["registrant_name"])
            ws.cell(row=row, column=4, value=reg["email"])
            ws.cell(row=row, column=5, value=reg["desired_table"] or "–")
            ws.cell(row=row, column=6, value=reg["total_tickets"])
            ws.cell(row=row, column=7, value=reg["total_tickets"] * price)
            ws.cell(row=row, column=8, value="✓ Ja" if reg["payment_confirmed"] else "✗ Nein")
            ws.cell(row=row, column=9, value=person["name"])
            ws.cell(row=row, column=10, value="Ja" if person["is_over_18"] else "Nein")
            ws.cell(
                row=row,
                column=11,
                value="Vegetarisch" if person["food_preference"] == "vegetarian" else "Fleisch",
            )
            ws.cell(
                row=row,
                column=12,
                value="Anmelder" if person["is_registrant"] else "Begleitung",
            )
            row += 1

    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(width + 4, 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=abbiball_registrierungen.xlsx"},
    )
